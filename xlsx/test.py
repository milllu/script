from openpyxl import load_workbook, Workbook
import re, json

def data_save(filename, dic):
    with open('data.txt', 'w') as f:
        f.write(json.dumps(dic))
    
def data_take(filename):
    with open('data.txt', 'r') as f:
        rt = json.loads(f.read())
    return rt

def excel_write(filename, ):
    wb = Workbook()
    sheet = wb.create_sheet('web', index=0)
    dic = data_take('data.txt')
    for key in dic:
        row = [key]
        for lst in dic[key]:
            row.extend(lst)
        sheet.append(row)
    wb.save('web_https.xlsx')

def excel(excelname, col1, col2, col3, dic):
    excel = load_workbook(filename=excelname)
    sheet1 = excel.sheetnames[0]
    table = excel[sheet1]
    rows = table.max_row

    for i in range(2, rows):
        data1 = table.cell(row=i, column=col1).value
        data2 = table.cell(row=i, column=col2).value
        data3 = table.cell(row=i, column=col3).value
        if data1:
            data1 = ''.join(data1.split())
            data1 = data1[:-1] if data1 and data1[-1] == '/' else data1
    
        if data1:
            if data1 not in dic:
                if data2:
                    dic[data1] = [[data2, data3]]
                else:
                    dic[data1] = []
            else:
                if data2 and [data2, data3] not in dic[data1]:
                    dic[data1].append([data2, data3])
    return dic


def main():
    dic = excel('Amy_web_login.xlsx', 2, 3, 4, {})
    dic2 = excel('Chris_web_login.xlsx', 2, 3, 4, dic)
    dic3 = excel('Jack_web_article_3.25.xlsx', 2, 3, 4, dic2)
    dic4 = excel('Jack_web_article.xlsx', 2, 3, 4, dic3)
    dic5 = excel('lily_uploads.xlsx', 2, 3, 4, dic4)
    dic6 = excel('may_login.xlsx', 2, 3, 4, dic5)
    dic7 = excel('web_login.xlsx', 2, 3, 4, dic6)

    return dic7
    
if __name__ == '__main__':
    # 保存
    # dic = main()
    # data_save('data.txt', dic)

    # 取出
    # excel_write('web_https.xlsx')
    dic = data_take('data.txt')
    print(dic)
    

    